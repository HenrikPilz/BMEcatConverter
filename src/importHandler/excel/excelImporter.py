'''
Created on 11.05.2017

@author: henrik.pilz
'''

import logging
import copy
import regex
from openpyxl import load_workbook
from multiprocessing.connection import arbitrary_address
from inspect import currentframe
from data.product import Product
from data.productDetails import ProductDetails
from data.orderDetails import OrderDetails
from data.featureSet import FeatureSet
from data.priceDetails import PriceDetails

class ExcelImporter(object):
    
    # Namen der mögliche Tabellen mit Artikeln
    __allowedTablenames = [ 'Artikel', 'Tabelle1', 'Mapping-Master' ]
    
    # Grunddaten eines Artikels. Mindestens die Artikelnummer
    basefieldMapping = {
        # Produktrumpf
        "supplierArticleId" : "productId"
    }
    
    # Artikeldetails
    productDetailMapping = {
        # Produktdetails
        "descriptionShort" : "title",
        "descriptionLong" : "description",
        "ean" : "ean",
        "supplierAltId" : "supplierAltId",
        "buyerId" : "buyerId",
        "manufacturerArticleId" : "manufacturerArticleId",
        "manufacturerName" : "manufacturerName",
        "deliveryTime" : "deliveryTime"
    }
    
    # Bestelldetails
    orderDetailMapping = {
        # Order Details
        "orderUnit" : "orderUnit",
        "contentUnit" : "contentUnit",
        "packingQuantity" : "packingQuantity",
        "priceQuantity" : "priceQuantity",
        "quantityMin" : "quantityMin",
        "quantityInterval" : "quantityInterval"
    }
    
    # technische Daten
    featureMapping = { "attributeName" : "name", "attributeValue" : "value" }
    
    # Preisdaten
    priceMapping = {
        "priceType" : "priceType",
        "priceAmount" : "amount",
        "tax" : "tax",
        "lowerBound" : "lowerBound",
        "factor" : "factor",
        "territory" : "territory",
        "currency" : "currency"
        }

    # Bilddaten
    mimeMapping = {
        "mimeType" : "mimeType",
        "mimeSource" : "source",
        "mimeDescription" : "description",
        "mimeAlt" : "altenativeContent",
        "mimePurpose" : "purpose",
        "mimeOrder" : "order"
        }

    '''
    classdocs
    '''
    def __init__(self):
        '''
        Constructor
        '''
        self.__indexForProduct = {}
        self.__indexForProductDetails = {}
        self.__indexForOrderDetails = {}
        self.__indexPairsForFeatures = {}
        self.__indexTuplesForPrices = {}
        self.__indexTuplesForMimes = {}
        self.articles = []
        
    def readWorkbook(self, filename):
        wb = load_workbook(filename)
        countPossibleCandidates = 0
        tablename = None
        for allowedSheetname in self.__allowedTablenames:
            if allowedSheetname in wb.sheetnames:
                countPossibleCandidates += 1
                tablename = allowedSheetname
        
        if tablename is None:
            raise Exception("Das Tabellenblatt mit den Artikelnamen sollte einen der folgenden Namen tragen: '{0}'".format(", ".join(self.__allowedTablenames)))
        if countPossibleCandidates > 1:
            raise Exception("Es darf nur ein Tabellenblatt mit den folgenden Artikelnamen existieren: '{0}'".format(", ".join(self.__allowedTablenames)))
        
        articleSheet = wb[tablename]
        self.__determineIndexMappings(articleSheet)
        self.__readArticles(articleSheet)
        
    def __determineIndexMappings(self, sheet):
        # gehe durch alle Spalten in Zeile 1 (Headerzeile)
        for colIndex in range(1,sheet.max_column):        
            # hole den aktuellen Feldnamen    
            currentFieldname = sheet.cell(column=colIndex, row=1).value
            # wenn der Feldnam nicht leer ist
            if currentFieldname is not None and currentFieldname.strip() is not None and len(currentFieldname.strip()) > 0:
                # gib ihn aus
                print (currentFieldname)
                # ist er in den basefieldKeys (Grunddatenfelder des Artikels)
                if currentFieldname in self.basefieldMapping.keys():
                    self.__indexForProduct[self.basefieldMapping[currentFieldname]] = colIndex
                # ist er product detail relevant
                elif currentFieldname in self.productDetailMapping.keys():                    
                    self.__indexForProductDetails[self.productDetailMapping[currentFieldname]] = colIndex
                # bestelldetails
                elif currentFieldname in self.orderDetailMapping.keys():
                    self.__indexForOrderDetails[self.orderDetailMapping[currentFieldname]] = colIndex
                # andere Daten?                
                else:
                    # dann muessen wir splitten, da der Zahlenanteil
                    # Sowohl die Zusammengehörigkeit der Daten anzeigt,
                    # als auch die Reihenfolge bestimmt
                    fieldName = regex.match("[a-zA-Z]*", currentFieldname).group(0)
                    fieldCount = currentFieldname.replace(fieldName, "")
                    
                    print("'{0}' : '{1}'".format(fieldName, fieldCount)) #logging.debug
                    
                    columnNameToClassFieldName = None
                    indexForClassFieldName = None
                    
                    # sind es preisdetails ?
                    if fieldName in self.priceMapping.keys():
                        indexForClassFieldName = self.__indexTuplesForPrices
                        columnNameToClassFieldName = self.priceMapping
                    elif fieldName in self.mimeMapping.keys():
                        indexForClassFieldName = self.__indexTuplesForMimes
                        columnNameToClassFieldName = self.mimeMapping
                    elif fieldName in self.featureMapping.keys():
                        indexForClassFieldName = self.__indexPairsForFeatures
                        columnNameToClassFieldName = self.featureMapping
                    else:
                        logging.debug("'{0}' : '{1}'".format(columnNameToClassFieldName[fieldName], fieldCount))
                        continue

                    classFieldName = columnNameToClassFieldName[fieldName]                        
                    if classFieldName not in indexForClassFieldName.keys():
                        indexForClassFieldName[classFieldName] = { fieldCount : colIndex } 
                    else:
                        indexForClassFieldName[classFieldName][fieldCount] = colIndex


    def __createProduct(self, sheet, rowIndex):
        currentProduct = Product()
        self.__transferInformationForMapping(self.__indexForProduct, currentProduct, rowIndex, sheet)
        currentProduct.details = ProductDetails()
        self.__transferInformationForMapping(self.__indexForProductDetails, currentProduct.details, rowIndex, sheet)
        currentProduct.orderDetails = OrderDetails()
        self.__transferInformationForMapping(self.__indexForOrderDetails, currentProduct.orderDetails, rowIndex, sheet)
        self.__addMultipleOrderedObjects(self.__indexTuplesForMimes, currentProduct, "Mime", rowIndex, sheet)
        priceDetails = PriceDetails()
        self.__addMultipleOrderedObjects(self.__indexPairsForFeatures, priceDetails, "Price", rowIndex, sheet)
        currentProduct.addPriceDetails(priceDetails)
        featureSet = FeatureSet()
        self.__addMultipleOrderedObjects(self.__indexPairsForFeatures, featureSet, "Feature", rowIndex, sheet)
        currentProduct.addFeatureSet(featureSet)
        return currentProduct

    def __readArticles(self, sheet):
        for rowIndex in range(2, sheet.max_row):
            currentProduct = self.__createProduct(sheet, rowIndex)
            self.articles.append(currentProduct)

    def __addMultipleOrderedObjects(self, mapping, objectToAddMultiplesTo, typeOfMultiplesAsString, rowIndex, sheet):
        typeOfMultiples = globals()[typeOfMultiplesAsString]
        itemsToAddByOrder = {}
        for fieldname in mapping.keys():
            for order, colIndex in mapping[fieldname].iteritems():
                if order not in itemsToAddByOrder.keys(): 
                    itemsToAddByOrder[order]= typeOfMultiples()                
                setattr(itemsToAddByOrder[order], fieldname, sheet.cell(columnIndex=colIndex, row=rowIndex))
        
        for key in sorted(itemsToAddByOrder.keys()):
            self.__exectueAddMethod(objectToAddMultiplesTo, "add" + str(typeOfMultiples), itemsToAddByOrder[key]) 

    def __transferInformationForMapping(self, mapping, objectForValue, rowIndex, sheet):
        for fieldname in mapping.keys():
            setattr(objectForValue, fieldname, sheet.cell(columnIndex=mapping[fieldname], row=rowIndex))
            
    def __exectueAddMethod(self, objectWithAddMethod ,addMethodName, arg):
        elementHandler = None
        try:
            if not addMethodName is None:
                elementHandler = getattr(objectWithAddMethod, addMethodName)
                elementHandler(arg)
            self.__currentContent = ""
        except AttributeError:
            raise NotImplementedError("Class [" + objectWithAddMethod.__class__.__name__ + "] does not implement [" + addMethodName + "]")
