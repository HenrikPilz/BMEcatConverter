"""
Created on 05.05.2017

@author: henrik.pilz

This module contains the Handler for the Excel output.
It's based on the data structures of the data package.

"""
import copy
import logging

from openpyxl.workbook import Workbook


class PyxelHandler(object):    
    """
    classdocs
    
    """
    
    baseFields = [ "articleType", "articleId", "supplierArticleId",
                   "descriptionShort", "descriptionLong", "ean",
                   "manufacturerArticleId", "manufacturerName",
                   "deliveryTime", "orderUnit",
                   "contentUnit", "packingQuantity", "priceQuantity",
                   "quantityMin", "quantityInterval" ]
    
    priceFields = [ "validFrom", "validTo", "priceType", "priceAmount",
                    "priceCurrency", "tax", "priceFactor", "lowerBound" ]
    
    mimeFields = [ "mimeType", "mimeSource", "mimeAlt", "mimeDescription",
                   "mimePurpose", "mimeOrder" ]
    
    attributeFields = [ "attributeName", "attributeValue" ] #, "attributeUnit" ]

    treatmentClassFields = [ "classType", "className" ]

    def __init__(self, articles, filename, defaultManufacturerName=None):
        """
        Constructor
        """
        self._workbook = None
        self._currentSheet = None        
        self._defaultManufacturerName = defaultManufacturerName
        self._headerRowIndex = 1
        self._filename = filename
        self._articles = copy.deepcopy(articles)
        self._firstPriceColumIndex = -1
        self._firstAttributeColumIndex = -1
        self._firstMimeColumIndex = -1
        self._firstSpecialTreatmentColumIndex = -1

        # Anzahl Preise, Attribute, Bilder, TreatmentClasses, Artikel 
        self._maxNumberOfPrices = 0
        self._maxNumberOfAttributes = 0
        self._maxNumberOfMimes = 0
        self._maxNumberOfSpecialTreatmentClasses = 0
        self._numberOfArticlesProcessed = 0

        self.inspectArticleSets()

        logging.info( "Anzahl verarbeiteter Artikel: " + str(self._numberOfArticlesProcessed))
        logging.info( "Maximale Anzahl Preise: " + str(self._maxNumberOfPrices))
        logging.info( "Maximale Anzahl Bilder: " + str(self._maxNumberOfMimes))
        logging.info( "Maximale Anzahl Attribute: " + str(self._maxNumberOfAttributes))
        logging.info( "Maximale Anzahl Spezialbehandlungsklassen: " + str(self._maxNumberOfSpecialTreatmentClasses))

    def __inspectArticleSets(self):
        """
         inspect Each set of articles 'new', 'update', 'delete' 
        """
        # For every entry articles (new, update, delete)
        for articles in self._articles.values():
            self.__inspectArticleSet(articles)

    def __inspectArticleSet(self, articles):
        """ inspect each article in one set of articles """
        # for each article
        for article in articles:
            # inspect article for data
            self.__inspectArticle(article)

    def __inspectArticle(self, article):
        """ 
        Determine the max occurrence of Mime-, Feature- and Price-entities.
        Determine the number variants this article describes.
        """
        # number of mimes
        self._maxNumberOfMimes = max(len(article.mimeInfo), self._maxNumberOfMimes)
        # number of special treatment classes
        self._maxNumberOfSpecialTreatmentClasses = max(len(article.details.specialTreatmentClasses), self._maxNumberOfSpecialTreatmentClasses)
        # number of prices
        numberOfPrices = self.__determinePriceCount(article.priceDetails)
        self._maxNumberOfPrices = max(numberOfPrices, self._maxNumberOfPrices)
        # number of variants and features
        numberOfVariants, numberOfAttributes = self.__determineFeatureCountAndVariantCount(article.featureSets)
        self._maxNumberOfAttributes = max(numberOfAttributes, self._maxNumberOfAttributes)
        article.numberOfVariants = numberOfVariants
        
        # number of processed articles
        self._numberOfArticlesProcessed += numberOfVariants


    def __determinePriceCount(self, priceDetails):
        """ Determine the max occurrence of the Price Entity"""
        numberOfPrices = 0
        for priceDetailEntry in priceDetails:
            numberOfPrices += len(priceDetailEntry.prices)
        
        return numberOfPrices

    def __determineVariantCountForFeatureSet(self, features):
        numberOfVariants = 1
        for feature in features:
            if feature.variants is not None:
                numberOfVariants = numberOfVariants * len(feature.variants)
        
        return numberOfVariants

    def __determineFeatureCountAndVariantCount(self, featureSets):
        numberOfVariants = 1
        numberOfAttributes = 0
        for featureSet in featureSets:
            numberOfVariantsForFeatureSet = self.__determineVariantCountForFeatureSet(featureSet.features)
            numberOfVariants = max(numberOfVariantsForFeatureSet, numberOfVariants)
            numberOfAttributes += len(featureSet.features)
        
        return numberOfVariants, numberOfAttributes

    def createNewWorkbookAndTransferData(self):
        self._workbook = Workbook()
        self.__createProductSheet()
        self._workbook.save(self._filename)
        self.__createReferencesSheet()
        self._workbook.save(self._filename)        
        self.__createKeywordsSheet()
        self._workbook.save(self._filename)

    def __createProductSheet(self):
        logging.info("Übertrage Artikel.")
        self._currentSheet = self._workbook.create_sheet("Artikel", 0)
        self.__createProductHeader()
        self.__writeArticleSetsToSheet()

    def __createProductHeader(self):
        """
        Create the header row for the article sheet
        """
        currentColumnIndex = self.__writeHeaderForFieldset(1, self.baseFields)

        if self._maxNumberOfPrices > 0:
            self._firstPriceColumIndex = currentColumnIndex
            currentColumnIndex = self.__writeHeaderForIndexedFieldset(self._maxNumberOfPrices, currentColumnIndex, self.priceFields)

        if self._maxNumberOfMimes> 0:
            self._firstMimeColumIndex = currentColumnIndex
            currentColumnIndex = self.__writeHeaderForIndexedFieldset(self._maxNumberOfMimes, currentColumnIndex, self.mimeFields)

        if self._maxNumberOfAttributes > 0:
            self._firstAttributeColumIndex = currentColumnIndex
            currentColumnIndex = self.__writeHeaderForIndexedFieldset(self._maxNumberOfAttributes, currentColumnIndex, self.attributeFields)

        if self._maxNumberOfSpecialTreatmentClasses > 0:
            self._firstSpecialTreatmentColumIndex = currentColumnIndex
            self.__writeHeaderForIndexedFieldset(self._maxNumberOfSpecialTreatmentClasses, currentColumnIndex, self.treatmentClassFields)

    def __writeHeaderForIndexedFieldset(self, maxCount, columnIndex, fields):
        if maxCount > 0:
            for index in range(1, maxCount+1):
                columnIndex = self.__writeHeaderForFieldset(columnIndex, fields, index)
        return columnIndex

    def __writeHeaderForFieldset(self, columnIndex, fields, index=None):
        for fieldName in fields:
            if index is not None:
                fieldName += str(index)
            self._currentSheet.cell(row=self._headerRowIndex, column=columnIndex, value=fieldName)
            columnIndex += 1
        
        return columnIndex

    def __writeArticleSetsToSheet(self):
        rowIndex = self._headerRowIndex + 1
        for articleType in self._articles.keys():
            self._currentArticleType = articleType
            self.__writeArticlesToSheet(rowIndex)


    def __writeArticlesToSheet(self, sheet, rowIndex):
        for article in self._articles[self._currentArticleType]:
            self.__writeOneArticleToRow(article, rowIndex)
            rowIndex += 1

    def __writeOneArticleToRow(self, article, rowIndex):
        self.__addBaseFieldsToArticle(article, rowIndex)
        self.__addMimesToArticle(article.mimeInfo, rowIndex)
        self.__addPricesToArticle(article.priceDetails, rowIndex)
        self.__addAttributesToArticle(article.featureSets, rowIndex)
        self.__addTreatmentClassesToArticle(article.details.specialTreatmentClasses, rowIndex)


    def __addBaseFieldsToArticle(self, article, rowIndex):
        columnIndex = 1
        self._currentSheet.cell(row=rowIndex, column=columnIndex, value=self._currentArticleType)
        columnIndex += 1
        self._currentSheet.cell(row=rowIndex, column=columnIndex, value="")
        columnIndex += 1
        self._currentSheet.cell(row=rowIndex, column=columnIndex, value=article.productId.strip().replace(" ",""))
        columnIndex += 1
        self._currentSheet.cell(row=rowIndex, column=columnIndex, value=article.details.title)
        columnIndex += 1
        self._currentSheet.cell(row=rowIndex, column=columnIndex, value=article.details.description)
        columnIndex += 1
        self._currentSheet.cell(row=rowIndex, column=columnIndex, value=article.details.ean)
        columnIndex += 1
        if article.details.manufacturerArticleId is None:
            self._currentSheet.cell(row=rowIndex, column=columnIndex, value=article.productId.strip())
        else:
            self._currentSheet.cell(row=rowIndex, column=columnIndex, value=article.details.manufacturerArticleId.strip())
        columnIndex += 1
        if article.details.manufacturerName is None:
            self._currentSheet.cell(row=rowIndex, column=columnIndex, value=self._defaultManufacturerName)
        else:
            self._currentSheet.cell(row=rowIndex, column=columnIndex, value=article.details.manufacturerName)
        columnIndex += 1
        self._currentSheet.cell(row=rowIndex, column=columnIndex, value=article.details.deliveryTime)
        columnIndex += 1
        self._currentSheet.cell(row=rowIndex, column=columnIndex, value=article.orderDetails.orderUnit)
        columnIndex += 1
        self._currentSheet.cell(row=rowIndex, column=columnIndex, value=article.orderDetails.contentUnit)
        columnIndex += 1
        self._currentSheet.cell(row=rowIndex, column=columnIndex, value=article.orderDetails.packingQuantity)
        columnIndex += 1
        self._currentSheet.cell(row=rowIndex, column=columnIndex, value=article.orderDetails.priceQuantity)
        columnIndex += 1
        self._currentSheet.cell(row=rowIndex, column=columnIndex, value=article.orderDetails.quantityMin)
        columnIndex += 1
        self._currentSheet.cell(row=rowIndex, column=columnIndex, value=article.orderDetails.quantityInterval)

    def __addPricesToArticle(self, priceDetails, rowIndex):
        if self._firstPriceColumIndex < 1:
            logging.warning("Keine Preise zu transferieren.")
            return
        priceCount = 0
        fieldCount = 8
        for priceDetail in priceDetails:
            for price in priceDetail.prices:
                self._currentSheet.cell(row=rowIndex, column=self._firstPriceColumIndex + priceCount*fieldCount, value=priceDetail.validFrom)
                self._currentSheet.cell(row=rowIndex, column=self._firstPriceColumIndex + priceCount*fieldCount + 1, value=priceDetail.validTo)
                self._currentSheet.cell(row=rowIndex, column=self._firstPriceColumIndex + priceCount*fieldCount + 2, value=price.priceType)
                self._currentSheet.cell(row=rowIndex, column=self._firstPriceColumIndex + priceCount*fieldCount + 3, value=price.amount)
                self._currentSheet.cell(row=rowIndex, column=self._firstPriceColumIndex + priceCount*fieldCount + 4, value=price.currency)
                self._currentSheet.cell(row=rowIndex, column=self._firstPriceColumIndex + priceCount*fieldCount + 5, value=price.tax)
                self._currentSheet.cell(row=rowIndex, column=self._firstPriceColumIndex + priceCount*fieldCount + 6, value=price.factor)
                self._currentSheet.cell(row=rowIndex, column=self._firstPriceColumIndex + priceCount*fieldCount + 7, value=price.lowerBound)
                priceCount += 1

    def __addMimesToArticle(self, mimeInfo, rowIndex):
        if self._firstMimeColumIndex < 1:
            logging.info("Keine Bilder zu transferieren.")
            return
        mimeCount = 0
        fieldCount = 6
        for mime in mimeInfo:
            self._currentSheet.cell(row=rowIndex, column=self._firstMimeColumIndex + mimeCount*fieldCount, value=mime.mimeType)
            self._currentSheet.cell(row=rowIndex, column=self._firstMimeColumIndex + mimeCount*fieldCount + 1, value=mime.source)
            self._currentSheet.cell(row=rowIndex, column=self._firstMimeColumIndex + mimeCount*fieldCount + 2, value=mime.altenativeContent)
            self._currentSheet.cell(row=rowIndex, column=self._firstMimeColumIndex + mimeCount*fieldCount + 3, value=mime.description)
            self._currentSheet.cell(row=rowIndex, column=self._firstMimeColumIndex + mimeCount*fieldCount + 4, value=mime.purpose)
            self._currentSheet.cell(row=rowIndex, column=self._firstMimeColumIndex + mimeCount*fieldCount + 5, value=mime.order)
            mimeCount += 1

    def __addTreatmentClassesToArticle(self, specialTreatmentClasses, rowIndex):
        if self._firstSpecialTreatmentColumIndex < 1:
            logging.info("Keine Fälle für Spezialbehandlungen zu transferieren.")
            return
        classCount = 0
        fieldCount = 2
        for treatmentClass in specialTreatmentClasses:
            self._currentSheet.cell(row=rowIndex, column=self._firstSpecialTreatmentColumIndex + classCount*fieldCount, value=treatmentClass.classType)
            self._currentSheet.cell(row=rowIndex, column=self._firstSpecialTreatmentColumIndex + classCount*fieldCount + 1, value=treatmentClass.value)
            classCount += 1

    def __addFieldValuesFromContainer(self, rowIndex, firstColumnIndex, fieldValueContainer, fieldNames):
        if firstColumnIndex < 1:
            logging.info("Keine Items zu transferieren.")
            return
        itemCount = 0
        fieldNameCount = len(fieldNames)
        for item in fieldValueContainer:
            fieldCount = 0
            for fieldName in fieldNameCount:
                self._currentSheet.cell(row=rowIndex, column=firstColumnIndex + itemCount*fieldNameCount + fieldCount, value=getattr(item, fieldName))
                fieldCount += 1
            itemCount += 1


    def __addAttributesToArticle(self, featureSets, rowIndex, currentVariant=None):
        if self._firstAttributeColumIndex < 1:
            logging.info("Keine Attribute zu transferieren.")
            return
        attributeCount = 0
        fieldCount = 2
        for featureSet in featureSets:
            for feature in featureSet.features:
                self._currentSheet.cell(row=rowIndex, column=self._firstAttributeColumIndex + attributeCount*fieldCount, value=feature.name)

                cellValue = None
                for value in feature.values:
                    if cellValue is None:
                        cellValue = value
                    else:
                        cellValue += " | "
                        cellValue += value 
                    
                    if not feature.unit is None and not feature.unit.strip() == "":
                        cellValue += " " + feature.unit

                self._currentSheet.cell(row=rowIndex, column=self._firstAttributeColumIndex + attributeCount*fieldCount + 1, value=cellValue)
                attributeCount += 1

    def __createReferencesSheet(self, wb):
        logging.info("Übertrage Artikelbeziehungen.")
        self._currentSheet = self._workbook .create_sheet("Artikelbeziehungen", 1)
        self.__createReferencesHeader()
        self.__writeReferencesToSheet()

    def __createReferencesHeader(self):
        columnIndex = 1
        self._currentSheet.cell(row=self._headerRowIndex, column=columnIndex, value="supplierArticleId")
        columnIndex += 1
        self._currentSheet.cell(row=self._headerRowIndex, column=columnIndex, value="referencType")
        columnIndex += 1
        self._currentSheet.cell(row=self._headerRowIndex, column=columnIndex, value="referencedSupplierArticleId")


    def __writeReferencesToSheet(self):
        rowIndex = self._headerRowIndex + 1
        for articleType in self._articles.keys():
            self._currentArticleType = articleType
            rowIndex += self.__writeReferencesForArticles(rowIndex)

    def __writeReferencesForArticles(self, rowIndex):
        currentRowIndex = rowIndex
        for article in self._articles[self._currentArticleType]:
            currentRowIndex += self.__writeReferencesForOneArticle(currentRowIndex, article)
        return currentRowIndex

    def __writeReferencesForOneArticle(self, rowIndex, article):
        """
        Transferiert alle gefunden Artikelbeziehungen für den aktuellen Artikel
        """
        # counts transferred references
        currentRowIndex = rowIndex
        for reference in article.references:
            currentRowIndex += self.__writeOneReferenceForOneArticle(currentRowIndex, article, reference)
        return currentRowIndex

    def __writeOneReferenceForOneArticle(self, rowIndex, article, reference):
        referencedArticleCount = 0
        for referencedArticleId in reference.supplierArticleIds:
            columnIndex = 1
            self._currentSheet.cell(row=rowIndex + referencedArticleCount, column=columnIndex, value=article.productId)
            columnIndex += 1
            self._currentSheet.cell(row=rowIndex + referencedArticleCount, column=columnIndex, value=reference.referenceType)
            columnIndex += 1
            self._currentSheet.cell(row=rowIndex + referencedArticleCount, column=columnIndex, value=referencedArticleId)
            referencedArticleCount += 1
        
        return referencedArticleCount

    def __createKeywordsSheet(self, wb):
        logging.info("Übertrage Artikelsuchbegriffe.")
        self._currentSheet = self._workbook.create_sheet("Artikelsuchbegriffe", 1)
        self.__createKeywordsHeader()
        self.__writeKeywordsForArticleSetsToSheet()

    def __createKeywordsHeader(self):
        columnIndex = 1
        self._currentSheet.cell(row=self._headerRowIndex, column=columnIndex, value="supplierArticleId")
        columnIndex += 1
        self._currentSheet.cell(row=self._headerRowIndex, column=columnIndex, value="keywords")

    def __writeKeywordsForArticleSetsToSheet(self):
        rowIndex = self._headerRowIndex + 1
        for articleType in self._articles.keys():
            self._currentArticleType = articleType
            self.__writeKeywordsForArticlesToSheet(rowIndex)

    def __writeKeywordsForArticlesToSheet(self, rowIndex):
        for article in self._articles[self._currentArticleType]:
            if len(article.details.keywords) > 0:
                self.__writeKeywordsForOneArticle(rowIndex, article)
                rowIndex += 1

    def __writeKeywordsForOneArticle(self, rowIndex, article):
        columnIndex = 1
        self._currentSheet.cell(row=rowIndex, column=columnIndex, value=article.productId)
        columnIndex += 1
        self._currentSheet.cell(row=rowIndex, column=columnIndex, value=",".join(['"{0}"'.format(keyword) for keyword in article.details.keywords]) )
