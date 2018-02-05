'''
Created on 11.05.2017

@author: henrik.pilz
'''

import logging

from openpyxl import load_workbook
import regex

from datamodel import Feature
from datamodel import FeatureSet
from datamodel import Mime
from datamodel import Price
from datamodel import PriceDetails
from datamodel import Product
from datamodel.validatingObject import FormulaFoundException
from transformer import NumberFormatException
from transformer import SeparatorTransformer


class ExcelImporter(object):
    '''
    classdocs
    '''

    # Namen der mögliche Tabellen mit Artikeln
    __allowedTablenames = [ 'Artikel', 'Tabelle1', 'Mapping-Master' ]

    # Grunddaten eines Artikels. Mindestens die Artikelnummer
    __basefieldMapping = {
        # Produktrumpf
        "supplierArticleId" : "productId"
    }

    # Artikeldetails
    __productDetailMapping = {
        # Produktdetails
        "descriptionShort" : "title",
        "descriptionLong" : "description",
        "ean" : "ean",
        "supplierAltId" : "supplierAltId",
        "buyerId" : "buyerId",
        "manufacturerArticleId" : "manufacturerArticleId",
        "manufacturerName" : "manufacturerName",
        "deliveryTime" : "deliveryTime",
        "delivery_time" : "deliveryTime"
    }

    # Bestelldetails
    __orderDetailMapping = {
        # Order Details
        "orderUnit" : "orderUnit",
        "contentUnit" : "contentUnit",
        "ContentUnit" : "contentUnit",
        "packingQuantity" : "packingQuantity",
        "priceQuantity" : "priceQuantity",
        "quantityMin" : "quantityMin",
        "quantityInterval" : "quantityInterval"
    }

    # technische Daten
    __featureMapping = {
        "attribute_name" : "name",
        "attributeName" : "name",
        "attribute_value" : "values",
        "attributeValue" : "values",
        "attribute_unit" : "unit",
        "attributeUnit" : "unit"
    }

    # Preisdaten
    __priceMapping = {
        "priceType" : "priceType",
        "price_type" : "priceType",
        "priceAmount" : "amount",
        "price_amount" : "amount",
        "tax" : "tax",
        "lowerBound" : "lowerBound",
        "lower_bound" : "lowerBound",
        "factor" : "factor",
        "territory" : "territory",
        "currency" : "currency"
    }

    # Bilddaten
    __mimeMapping = {
        "mimeType" : "mimeType",
        "mime_type" : "mimeType",
        "mimeSource" : "source",
        "mime_source" : "source",
        "mimeDescription" : "description",
        "mime_description" : "description",
        "mimeAlt" : "altenativeContent",
        "mime_alt" : "altenativeContent",
        "mimePurpose" : "purpose",
        "mime_purpose" : "purpose",
        "mimeOrder" : "order",
        "mime_order" : "order"
    }

    __fieldsToTransform = [ "priceAmount", "price_amount", "amount", "tax", "factor", "delivery_time", "deliveryTime" ]

    def __init__(self, separatorTransformer=SeparatorTransformer("detect")):
        '''
        Constructor
        '''
        self.__indexForProduct = {}
        self.__indexForProductDetails = {}
        self.__indexForOrderDetails = {}
        self.__indexPairsForFeatures = {}
        self.__indexTuplesForPrices = {}
        self.__indexTuplesForMimes = {}
        self.__currentRowIndex = 1
        self.__currentSheet = None
        self._separatorTransformer = separatorTransformer

        self.articles = []

    def readWorkbook(self, filename):
        wb = load_workbook(filename)
        countPossibleCandidates = 0
        tablename = None
        for allowedSheetname in self.__allowedTablenames:
            if allowedSheetname in wb.sheetnames:
                countPossibleCandidates += 1
                tablename = allowedSheetname

        if tablename is None:
            raise Exception("Das Tabellenblatt mit den Artikelnamen sollte einen der folgenden Namen tragen: '{0}'".format(", ".join(self.__allowedTablenames)))
        if countPossibleCandidates > 1:
            raise Exception("Es darf nur ein Tabellenblatt mit den folgenden Artikelnamen existieren: '{0}'".format(", ".join(self.__allowedTablenames)))

        self.__currentSheet = wb[tablename]
        self.__determineIndexMappings()
        self.__readArticles()

    def __determineIndexMappings(self):
        # gehe durch alle Spalten in Zeile 1 (Headerzeile)
        for colIndex in range(1, self.__currentSheet.max_column + 1):
            # hole den aktuellen Feldnamen
            currentFieldname = self.__currentSheet.cell(column=colIndex, row=1).value
            # wenn der Feldnam nicht leer ist
            if currentFieldname is not None and len(currentFieldname.strip()) > 0:
                self.__detectEntities(currentFieldname, colIndex)

    def __detectEntities(self, currentFieldname, colIndex):
        # gib ihn aus
        logging.debug(currentFieldname)

        # ist er in den basefieldKeys (Grunddatenfelder des Artikels)
        added = self.__addIndexForMappingIfFieldInMapping(currentFieldname, colIndex, self.__basefieldMapping, self.__indexForProduct)
        # ist er product detail relevant
        added = added or self.__addIndexForMappingIfFieldInMapping(currentFieldname, colIndex, self.__productDetailMapping, self.__indexForProductDetails)
        # bestelldetails
        added = added or self.__addIndexForMappingIfFieldInMapping(currentFieldname, colIndex, self.__orderDetailMapping, self.__indexForOrderDetails)

        # andere Daten?
        if not added:
            self.__detectMultiColumnEntities(currentFieldname, colIndex)

    def __addIndexForMappingIfFieldInMapping(self, currentFieldname, colIndex, mappingDictionary, indexDictionary):
        # ist er in den basefieldKeys (Grunddatenfelder des Artikels)
        if currentFieldname in mappingDictionary.keys():
            indexDictionary[mappingDictionary[currentFieldname]] = colIndex
            return True
        return False

    def __detectMultiColumnEntities(self, currentFieldname, colIndex):
        # dann muessen wir splitten, da der Zahlenanteil
        # Sowohl die Zusammengehörigkeit der Daten anzeigt,
        # als auch die Reihenfolge bestimmt
        fieldName = regex.match("[a-zA-Z_]*", currentFieldname).group(0)
        fieldName = fieldName.strip("_")
        fieldCount = currentFieldname.replace(fieldName, "").strip("_")

        logging.debug("'{0}' : '{1}'".format(fieldName, fieldCount))

        # sind es Preise ?
        added = self.__addIndexForFieldCountMappingIfFieldInMapping(fieldName, colIndex, fieldCount,
                                                                    self.__priceMapping, self.__indexTuplesForPrices)
        # sind es Bilder ?
        added = added or self.__addIndexForFieldCountMappingIfFieldInMapping(fieldName, colIndex, fieldCount,
                                                                             self.__mimeMapping, self.__indexTuplesForMimes)
        # sind es Attribute
        added = added or self.__addIndexForFieldCountMappingIfFieldInMapping(fieldName, colIndex, fieldCount,
                                                                             self.__featureMapping, self.__indexPairsForFeatures)

        if not added:
            logging.debug("'{0}' : '{1}'".format(fieldName, fieldCount))

    def __addIndexForFieldCountMappingIfFieldInMapping(self, currentFieldname, colIndex, fieldCount, mappingDictionary, indexDictionary):
        if currentFieldname in mappingDictionary.keys():
            self.__setColumnIndexForMapping(indexDictionary, mappingDictionary[currentFieldname], fieldCount, colIndex)
            return True
        return False

    def __setColumnIndexForMapping(self, indexForClassFieldName, classFieldName, fieldCount, colIndex):
        if classFieldName not in indexForClassFieldName.keys():
            indexForClassFieldName[classFieldName] = { fieldCount : colIndex }
        else:
            indexForClassFieldName[classFieldName][fieldCount] = colIndex

    def __readArticles(self):
        for rowIndex in range(2, self.__currentSheet.max_row + 1):
            self.__currentRowIndex = rowIndex
            self.articles.append(self.__createProduct())

    def __createProduct(self):
        currentProduct = Product()
        self.__transferInformationForMapping(self.__indexForProduct, currentProduct)
        currentProduct.addDetails()
        try:
            self.__transferInformationForMapping(self.__indexForProductDetails, currentProduct.details)
            currentProduct.addOrderDetails()
            self.__transferInformationForMapping(self.__indexForOrderDetails, currentProduct.orderDetails)
            self.__addMultipleOrderedObjects(self.__indexTuplesForMimes, currentProduct, Mime)
            priceDetails = PriceDetails()
            self.__addMultipleOrderedObjects(self.__indexTuplesForPrices, priceDetails, Price)
            currentProduct.addPriceDetails(priceDetails, raiseException=False)
            featureSet = FeatureSet()
            self.__addMultipleOrderedObjects(self.__indexPairsForFeatures, featureSet, Feature)
            currentProduct.addFeatureSet(featureSet)
            currentProduct.validate(raiseException=False)
        except FormulaFoundException as ffe:
            raise FormulaFoundException("{0}: {1}".format(currentProduct.productId, str(ffe)))
        return currentProduct

    def __transferInformationForMapping(self, mapping, objectForValue):
        """
        Überträgt die Informationen vom Objekt über das angegebene Mapping für das Produkt in der Zeile.
        """
        for fieldname, colIndex in mapping.items():
            try:
                self.__determineAndAddValue(colIndex, objectForValue, fieldname)
            except NumberFormatException as e:
                raise NumberFormatException("Zeile: {0}/Spalte {1}; '{2}' Fehler: {3}".format(self.__currentRowIndex, colIndex,
                                                                                              fieldname, str(e)))

    def __addMultipleOrderedObjects(self, mapping, objectContainer, typeOfMultiples):
        """
        Fügt mehrere Objekte zum Objektcontainer hinzu
        """
        itemsToAddByOrder = {}
        for fieldname in mapping.keys():
            self.__setValueForFieldNameAndTransformDecimalsIfNeeded(mapping, typeOfMultiples, itemsToAddByOrder, fieldname)

        self.__sortAndAddMultipleOrderedObjects(objectContainer, typeOfMultiples, itemsToAddByOrder)

    def __setValueForFieldNameAndTransformDecimalsIfNeeded(self, mapping, typeOfMultiples, itemsToAddByOrder, fieldname):
        for order, colIndex in mapping[fieldname].items():
            try:
                if order not in itemsToAddByOrder.keys():
                    itemsToAddByOrder[order] = typeOfMultiples()
                self.__determineAndAddValue(colIndex, itemsToAddByOrder[order], fieldname)
            except NumberFormatException as e:
                raise NumberFormatException("Zeile: {0}/Spalte {1}; '{2}{4}' Fehler: {3}".format(self.__currentRowIndex, colIndex,
                                                                                                 fieldname, str(e), order))

    def __determineAndAddValue(self, colIndex, objectForValue, fieldname):
        value = self.__currentSheet.cell(column=colIndex, row=self.__currentRowIndex).value
        value = self.__transformValueIfSupposedTo(fieldname, value)
        if value is not None and len(str(value)) > 0:
            objectForValue.add(fieldname, value)

    def __transformValueIfSupposedTo(self, fieldname, value):
        if fieldname in self.__fieldsToTransform:
            return self._separatorTransformer.transform(value)
        return value

    def __sortAndAddMultipleOrderedObjects(self, objectContainer, typeOfMultiples, itemsToAddByOrder):
        for key in sorted(itemsToAddByOrder.keys()):
            item = itemsToAddByOrder[key]
            self.__setOrderIfNotSet(item, key)
            self.__exectueAddMethod(objectContainer, "add" + str(typeOfMultiples.__name__), item)

    def __setOrderIfNotSet(self, objectWithOrder, order):
        try:
            if getattr(objectWithOrder, "order") is None:
                setattr(objectWithOrder, "order", int(order))
        except AttributeError:
            logging.debug("Order Attribute could not be set for '{0}'.".format(objectWithOrder.__class__.__name__))

    def __exectueAddMethod(self, objectWithAddMethod , addMethodName, arg):
        try:
            elementHandler = getattr(objectWithAddMethod, addMethodName)
            elementHandler(arg)
        except AttributeError:
            raise NotImplementedError("Class [" + objectWithAddMethod.__class__.__name__ + "] does not implement [" + addMethodName + "]")
