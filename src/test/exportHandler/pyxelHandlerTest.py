'''
Created on 08.10.2017

@author: Henrik Pilz
'''
import unittest
from exportHandler.excel.pyxelHandler import PyxelHandler
from data import Product, ProductDetails, TreatmentClass, Mime, Reference, OrderDetails, Feature, Price, PriceDetails
from openpyxl import load_workbook
from importHandler.excel.excelImporter import ExcelImporter

class PyxelHandlerTest(unittest.TestCase):


    def testCreateWorkbook(self):
        article = Product()
        article.productId = '12345'
        article.details = ProductDetails()
        article.details.deliveryTime = 10
        article.details.description = 'Test Description'
        article.details.ean = '12345678901234'
        article.details.keywords = [ 'Keyword 1', 'Keyword 2']
        article.details.manufacturerArticleId = '09876'
        article.details.manufacturerName = 'Manufacturer'
        tc = TreatmentClass()
        tc.classType = 'TestClass'
        tc.value = '12345'
        article.details.specialTreatmentClasses = [ tc ]
        article.details.title = 'Test Article'
        article.details.supplierAltId = '23456'
        reference = Reference()
        reference.referenceType = 'accessory'
        reference.supplierArticleIds = [ '09876', '7654' ]        
        article.addReference(reference)
        
        mime = Mime()
        mime.mimeType = 'image/jpg'
        mime.order = 1
        mime.purpose = 'detail'
        mime.source = 'manufacturer/Test.jpg'
        article.addMime(mime)
        article.orderDetails = OrderDetails()
        article.orderDetails.contentUnit = 'C62'
        article.orderDetails.orderUnit = 'C62'
        article.orderDetails.packingQuantity = 25
        article.orderDetails.priceQuantity = 100
        article.orderDetails.quantityMin = 4
        article.orderDetails.quantityInterval = 1
        
        priceDetails = PriceDetails()
        price = Price()
        price.amount = 10.50
        price.priceType ='net_customer'
        price.lowerBound = 1
        price.tax = 0.19
        priceDetails.addPrice(price)
        article.addPriceDetails(priceDetails)
        article.validate(True)
        
        articles = { 'new' : [ article ]}
        
        pyxelHandler = PyxelHandler(articles, 'test_excel_output.xlsx')
        
        pyxelHandler.createNewWorkbook()


        excelImporter = ExcelImporter()
        excelImporter.readWorkbook('test_excel_output.xlsx')
        
        article2 = excelImporter.articles[0]
        
        self.assertEqual(article, article2, "Test")
        
        self.assertEqual(article.productId, article2.productId, "Product ID not Equal")
        article.details = ProductDetails()
        article.details.deliveryTime = 10
        article.details.description = 'Test Description'
        article.details.ean = '12345678901234'
        article.details.keywords = [ 'Keyword 1', 'Keyword 2']
        article.details.manufacturerArticleId = '09876'
        article.details.manufacturerName = 'Manufacturer'
        tc = TreatmentClass()
        tc.classType = 'TestClass'
        tc.value = '12345'
        article.details.specialTreatmentClasses = [ tc ]
        article.details.title = 'Test Article'
        article.details.supplierAltId = '23456'
        reference = Reference()
        reference.referenceType = 'accessory'
        reference.supplierArticleIds = [ '09876', '7654' ]
        
        
if __name__ == "__main__":
    #import sys;sys.argv = ['', 'Test.testInit']
    unittest.main()