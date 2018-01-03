'''
Created on 05.05.2017

@author: henrik.pilz
'''
import copy
import logging

from openpyxl.workbook import Workbook


class PyxelExporter(object):
    '''
    classdocs
    '''

    __additionalSheetsMapping = [ ("Artikelbeziehungen", 1,
                                   [ "supplierArticleId", "referencType", "referencedSupplierArticleId" ],
                                   "_writeReferencesForOneArticle"),
                                  ("Artikelsuchbegriffe", 2,
                                   [ "supplierArticleId", "keywords" ],
                                   "_writeKeywordsForOneArticle")
                                  ]

    __baseFields = [ "articleType", "articleId", "supplierArticleId",
                     "descriptionShort", "descriptionLong", "ean",
                     "manufacturerArticleId", "manufacturerName",
                     "deliveryTime", "orderUnit",
                     "contentUnit", "packingQuantity", "priceQuantity",
                     "quantityMin", "quantityInterval" ]

    __priceFields = [ "validFrom", "validTo", "priceType", "priceAmount",
                      "priceCurrency", "tax", "priceFactor", "lowerBound" ]

    __mimeFields = [ "mimeType", "mimeSource",  # "mimeAlt",
                     "mimeDescription", "mimePurpose", "mimeOrder" ]

    __attributeFields = [ "attributeName", "attributeValue" ]  # , "attributeUnit" ]

    __treatmentClassFields = [ "classType", "className" ]

    __headerRowIndex = 1

    def __init__(self, articles, filename, defaultManufacturerName=None):
        '''
        Constructor
        '''
        self._filename = filename
        self._workbook = None
        self._defaultManufacturerName = defaultManufacturerName
        self._maxNumberOfPrices = None
        self._currentColumnIndex = 0
        self._currentRowIndex = 0
        self._articles = copy.deepcopy(articles)
        self._firstPriceColumIndex = -1
        self._firstAttributeColumIndex = -1
        self._firstMimeColumIndex = -1
        self._firstSpecialTreatmentColumIndex = -1
        ''' Anzahl Preise, Attribute, Bilder, TreatmentClasses, Artikel'''
        self._maxNumberOfPrices = 0
        self._maxNumberOfAttributes = 0
        self._maxNumberOfMimes = 0
        self._maxNumberOfSpecialTreatmentClasses = 0
        self._numberOfArticlesProcessed = 0
        self._numberOfArticlesWithKeywords = 0
        self._numberOfArticlereferences = 0

        for articleSet in self._articles.values():
            self.__countValuesForArticleSet(articleSet)

    def __countValuesForArticleSet(self, articleSet):
        for article in articleSet:
            self.__extractNumbersFromArticle(article)

    def __extractNumbersFromArticle(self, article):
        self._maxNumberOfPrices = max(self._maxNumberOfPrices, self.__extractCount(article.priceDetails))
        self._maxNumberOfAttributes = max(self._maxNumberOfAttributes, self.__extractCount(article.featureSets))
        self._numberOfArticlesProcessed += article.numberOfVariants
        self._maxNumberOfMimes = max(self._maxNumberOfMimes, len(article.mimeInfo))
        self._maxNumberOfSpecialTreatmentClasses = max(self._maxNumberOfSpecialTreatmentClasses, len(article.details.specialTreatmentClasses))

    def __extractCount(self, entryList):
        countEntries = 0
        for entry in entryList:
            countEntries += len(entry)

        return countEntries

    ''' ExcelMappe erstellen'''
    def createNewWorkbook(self):
        self._workbook = Workbook()
        self.__createArtikelSheet()
        self._workbook.save(self._filename)
        self.__createAdditionalSheets()
        self._workbook.save(self._filename)
        logging.info("Anzahl zu verarbeitender Artikel: " + str(self._numberOfArticlesProcessed))
        logging.info("Maximale Anzahl Preise: " + str(self._maxNumberOfPrices))
        logging.info("Maximale Anzahl Bilder: " + str(self._maxNumberOfMimes))
        logging.info("Maximale Anzahl Attribute: " + str(self._maxNumberOfAttributes))
        logging.info("Maximale Anzahl Spezialbehandlungsklassen: " + str(self._maxNumberOfSpecialTreatmentClasses))
        logging.info("Anzahl Artikelreferenzen: " + str(self._numberOfArticlereferences))
        logging.info("Anzahl Artikel mit Suchworteinträgen: " + str(self._numberOfArticlesWithKeywords))

    def __createArtikelSheet(self):
        logging.info("Übertrage Artikel.")
        self._currentSheet = self._workbook.create_sheet("Artikel", 0)
        self.__createArtikelHeader()
        self.__writeArticlesToSheet()

    ''' Artikelheader '''
    def __createArtikelHeader(self):
        self._currentColumnIndex = 1
        self._currentRowIndex = self.__headerRowIndex

        for fieldName in self.__baseFields:
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(fieldName)

        self._firstPriceColumIndex = self.__addArticlePartIfNecessary(self._maxNumberOfPrices, self.__priceFields)
        self._firstMimeColumIndex = self.__addArticlePartIfNecessary(self._maxNumberOfMimes, self.__mimeFields)
        self._firstAttributeColumIndex = self.__addArticlePartIfNecessary(self._maxNumberOfAttributes, self.__attributeFields)
        self._firstSpecialTreatmentColumIndex = self.__addArticlePartIfNecessary(self._maxNumberOfSpecialTreatmentClasses, self.__treatmentClassFields)

    def __addArticlePartIfNecessary(self, maxCount, entryList):
        startingColumnIndex = 0
        if maxCount > 0:
            startingColumnIndex = self._currentColumnIndex
            self.__iterateRangeForMaxCount(maxCount, entryList)

        return startingColumnIndex

    def __iterateRangeForMaxCount(self, maxCount, entryList):
        for i in range(1, maxCount + 1):
            self.__addFieldsetForCurrentCount(entryList, i)

    def __addFieldsetForCurrentCount(self, entryList, i):
        for fieldName in entryList:
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(fieldName + str(i))

    ''' Artikeldaten transferieren '''
    def __writeArticlesToSheet(self):
        self._currentRowIndex = self.__headerRowIndex + 1
        for articleType, articles in self._articles.items():
            self.__transferArticleSet(articleType, articles)

    def __transferArticleSet(self, articleType, articles):
        for article in articles:
            self.__writeOneArticleToRow(articleType, article)

    def __writeOneArticleToRow(self, articleType, article):
        self.__addBaseFieldsToArticle(articleType, article)
        self.__addMimesToArticle(article.mimeInfo)
        self.__addPriceDetailsToArticle(article.priceDetails)
        self.__addAttributesToArticle(article.featureSets)
        self.__addTreatmentClassesToArticle(article.details.specialTreatmentClasses)
        self._currentRowIndex += 1

    def __addBaseFieldsToArticle(self, articleType, article):
        self._currentColumnIndex = 1
        self.__writeValueToCurrentCellAndIncreaseColumnIndex(articleType)
        self.__writeValueToCurrentCellAndIncreaseColumnIndex("")
        self.__writeValueToCurrentCellAndIncreaseColumnIndex(article.productId.strip())
        self.__writeValueToCurrentCellAndIncreaseColumnIndex(article.details.title)
        self.__writeValueToCurrentCellAndIncreaseColumnIndex(article.details.description)
        self.__writeValueToCurrentCellAndIncreaseColumnIndex(article.details.ean)

        if article.details.manufacturerArticleId is None:
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(article.productId.strip())
        else:
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(article.details.manufacturerArticleId.strip())

        if article.details.manufacturerName is None:
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(self._defaultManufacturerName)
        else:
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(article.details.manufacturerName)

        self.__writeValueToCurrentCellAndIncreaseColumnIndex(article.details.deliveryTime)
        self.__writeValueToCurrentCellAndIncreaseColumnIndex(article.orderDetails.orderUnit)
        self.__writeValueToCurrentCellAndIncreaseColumnIndex(article.orderDetails.contentUnit)
        self.__writeValueToCurrentCellAndIncreaseColumnIndex(article.orderDetails.packingQuantity)
        self.__writeValueToCurrentCellAndIncreaseColumnIndex(article.orderDetails.priceQuantity)
        self.__writeValueToCurrentCellAndIncreaseColumnIndex(article.orderDetails.quantityMin)
        self.__writeValueToCurrentCellAndIncreaseColumnIndex(article.orderDetails.quantityInterval)

    def __transferPriceDetailForArticle(self, priceDetail):
        for price in priceDetail.prices:
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(priceDetail.validFrom)
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(priceDetail.validTo)
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(price.priceType)
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(price.amount)
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(price.currency)
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(price.tax)
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(price.factor)
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(price.lowerBound)

    def __addPriceDetailsToArticle(self, priceDetails):
        if self._firstPriceColumIndex < 1:
            logging.warning("Keine Preise zu transferieren.")
            return

        self._currentColumnIndex = self._firstPriceColumIndex

        for priceDetail in priceDetails:
            self.__transferPriceDetailForArticle(priceDetail)

    def __addMimesToArticle(self, mimeInfo):
        if self._firstMimeColumIndex < 1:
            logging.warning("Keine Bilder zu transferieren.")
            return
        self._currentColumnIndex = self._firstMimeColumIndex
        for mime in mimeInfo:
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(mime.mimeType)
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(mime.source)
            # self.__writeValueToCurrentCellAndIncreaseColumnIndex(mime.altenativeContent)
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(mime.description)
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(mime.purpose)
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(mime.order)

    def __addAttributesToArticle(self, featureSets, currentVariant=None):
        if self._firstAttributeColumIndex < 1:
            logging.warning("Keine Attribute zu transferieren.")
            return
        self._currentColumnIndex = self._firstAttributeColumIndex
        for featureSet in featureSets:
            self.__writeFeatureSetForArticle(featureSet.features)

    def __writeFeatureSetForArticle(self, features):
        for feature in features:
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(feature.name)
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(self.__getValuesFromFeature(feature))

    def __getValuesFromFeature(self, feature):
        cellValue = ""
        for value in feature.values:
            if len(cellValue) > 0:
                cellValue += " | "
            cellValue += self.__joinValueAndUnitIfNecessary(value, feature.unit)
        return cellValue

    def __joinValueAndUnitIfNecessary(self, value, unit):
        if unit is not None and unit.strip() != "":
            value = "{0} {1}".format(value, unit)
        return str(value)

    def __addTreatmentClassesToArticle(self, specialTreatmentClasses):
        if self._firstSpecialTreatmentColumIndex < 1:
            logging.debug("Keine Fälle für Spezialbehandlungen zu transferieren.")
            return
        self._currentColumnIndex = self._firstSpecialTreatmentColumIndex
        for treatmentClass in specialTreatmentClasses:
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(treatmentClass.classType)
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(treatmentClass.value)

    ''' create additional sheets '''
    def __createAdditionalSheets(self):
        for sheetName, index, columnNames, dataTransferMethodName in self.__additionalSheetsMapping:
            logging.info("Übertrage {0}".format(sheetName))
            self._currentSheet = self._workbook.create_sheet(sheetName, index)
            self.__createAdditionalSheetHeader(sheetName, columnNames)
            self.__writeAdditionalDataToSheet(dataTransferMethodName)

    ''' create additional sheets - Header '''
    def __createAdditionalSheetHeader(self, sheetName, additionalSheetMapping):
        self._currentColumnIndex = 1
        self._currentRowIndex = self.__headerRowIndex
        for columnName in additionalSheetMapping:
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(columnName)

    ''' create additional sheets - Transferdata '''
    def __writeAdditionalDataToSheet(self, dataTransferMethodName):
        self._currentRowIndex = self.__headerRowIndex + 1
        for articles in self._articles.values():
            self.__writeAdditionalDataForArticleSet(articles, dataTransferMethodName)

    def __writeAdditionalDataForArticleSet(self, articles, dataTransferMethodName):
        for article in articles:
            dataTransferMethod = getattr(self, dataTransferMethodName)
            dataTransferMethod(article)

    ''' Datetransfermethods '''
    def _writeReferencesForOneArticle(self, article):
        for reference in article.references:
            self._currentColumnIndex = 1
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(article.productId)
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(reference.referenceType)
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(reference.supplierArticleId)
            self._currentRowIndex += 1
            self._numberOfArticlereferences += 1

    def _writeKeywordsForOneArticle(self, article):
        self._currentColumnIndex = 1
        if len(article.details.keywords) > 0:
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(article.productId)
            self.__writeValueToCurrentCellAndIncreaseColumnIndex(",".join(['"{0}"'.format(keyword) for keyword in article.details.keywords]))
            self._currentRowIndex += 1
            self._numberOfArticlesWithKeywords += 1

    ''' Basic write to cell '''
    def __writeValueToCurrentCellAndIncreaseColumnIndex(self, valueToWrite):
        self._currentSheet.cell(row=self._currentRowIndex, column=self._currentColumnIndex, value=valueToWrite)
        self._currentColumnIndex += 1
